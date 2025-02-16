

# New File
# from openpyxl import Workbook

# wb = Workbook()

# sheet = wb.create_sheet(title='tab1')

# wb.save('New_tab.xlsx')

# Existing File
from openpyxl import load_workbook

wb2 = load_workbook('New_tab.xlsx')
sheet2 = wb2.create_sheet(title='tab2',
                          index=0)
wb2.save('New_tab.xlsx')


