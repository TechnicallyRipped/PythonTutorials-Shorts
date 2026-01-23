
from openpyxl import load_workbook

wb = load_workbook('scores_.xlsx')
ws = wb['Sheet1']

ws['A1'].hyperlink = 'https://www.google.com/'
ws['A1'].style = 'Hyperlink'

wb.save('hyperlink.xlsx')