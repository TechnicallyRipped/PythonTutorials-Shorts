


from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook('Example.xlsx')
ws = wb['Colors']


ws['A2'].fill = PatternFill(start_color='FF0000',
                             fill_type='solid')

ws['A3'].fill = PatternFill(start_color='FF0000',
                             end_color='FFBF00',
                               fill_type='lightVertical')

ws['A4'].fill = PatternFill(start_color='800080',
                            end_color='008000',
                            fill_type='lightGrid')

wb.save('Example.xlsx')











