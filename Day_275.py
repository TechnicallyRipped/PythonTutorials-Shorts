

from openpyxl import load_workbook
from openpyxl.styles import Font


wb = load_workbook('Example.xlsx')
ws = wb['Font']

ws['A1'].font = Font(name='Arial Nova',
                     size=20,
                     bold=True,
                     italic=True,
                     underline='single',
                     color='FF0000') #Hex value
ws['A2'].font = Font(size=30,
                     strike=True)

wb.save('Example.xlsx')
























