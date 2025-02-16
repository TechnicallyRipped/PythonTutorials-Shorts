


# pip install openpyxl

import openpyxl as xl

workbook = xl.load_workbook('Example.xlsx')
ws = workbook['Ages']

ws['A7'].value = 'Ian'
ws['B7'].value = '30'

workbook.save('Example.xlsx')

















