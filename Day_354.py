


from openpyxl import load_workbook

wb = load_workbook('merge_test.xlsx')

ws = wb['Sheet1']

ws.merge_cells('A1:B2')

ws['A1'] = 'Merged Cell'

wb.save('merge_test2.xlsx')






