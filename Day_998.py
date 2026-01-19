
from openpyxl import load_workbook

wb = load_workbook('scores_.xlsx')

ws = wb['Sheet1']

ws['B6'] = '=SUM(B2:B5)'

wb.save('scores_2.xlsx')