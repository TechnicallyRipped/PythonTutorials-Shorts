

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

wb = load_workbook('merge_test.xlsx')

ws = wb['Sheet1']

x = 'FF0000'

ws['B2'].border = Border(left=Side(style='thin',color=x),
                         right=Side(style='medium',color=x),
                         top=Side(style='dashed',color=x), 
                         bottom=Side(style='thick',color=x))

wb.save('borders.xlsx')